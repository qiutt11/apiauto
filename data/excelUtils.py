# -!- coding:utf-8 -!-
from openpyxl import load_workbook
import os

class doExcel():
    def __init__(self,filename=None,sheetId=None):

        if filename:
            self.filename=filename
            self.sheetId=sheetId
        else:
            self.filename="data.xlsx"
            self.sheetId="data"
        self.path = os.path.join(os.path.abspath(os.path.dirname(__file__)), self.filename)
        self.data=self.getData()
        print("数据来源：%s" % self.path)
    def getData(self):
        wb = load_workbook(self.path)
        sheet = wb[self.sheetId]

        test_data = []
        for i in range(1, sheet.max_row):
            sub_data = {}
            sub_data['caseId'] = sheet.cell(i + 1, 1).value
            sub_data['interface'] = sheet.cell(i + 1, 2).value
            sub_data['url'] = sheet.cell(i + 1, 3).value
            sub_data['isRun'] = sheet.cell(i + 1, 4).value
            sub_data['method'] = sheet.cell(i + 1, 5).value
            sub_data['headers'] = sheet.cell(i + 1, 6).value
            if sheet.cell(i + 1, 7).value=='':
                sub_data['header_depend']=''
            else:
                sub_data['header_depend'] = sheet.cell(i + 1, 7).value
            if  sheet.cell(i + 1, 8).value=='':
                sub_data['header_key'] == ''
            else:
                sub_data['header_key'] = sheet.cell(i + 1, 8).value
            sub_data['data'] = sheet.cell(i + 1, 9).value
            if sheet.cell(i + 1, 10).value=='':
                sub_data['data_depend']==''
            else:
                sub_data['data_depend'] = sheet.cell(i + 1, 10).value
            if sheet.cell(i + 1, 11).value=='':
                sub_data['data_key']==''
            else:
                sub_data['data_key'] = sheet.cell(i + 1, 11).value
            sub_data['expected'] = sheet.cell(i + 1, 12).value
            sub_data['check_type'] = sheet.cell(i + 1, 13).value  # 验证目的
            sub_data['actually'] = sheet.cell(i + 1, 14).value
            sub_data['aduit'] = sheet.cell(i + 1, 15).value
            sub_data['workflowid'] = sheet.cell(i + 1, 16).value
            #sub_data['sql'] = sheet.cell(i + 1, 17).value
            test_data.append(sub_data)
        return test_data

if __name__=='__main__':
    a=doExcel().getData()
    print(a)
